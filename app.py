"""
E-E-A-T Checker — Streamlit app
"""
import re
from io import BytesIO
from urllib.parse import urlparse

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, Alignment, PatternFill

from checkers import (
    OK, FAIL, WARN,
    fetch, get_internal_links, find_page_url, get_schemas, has_schema_type,
    chk_https, chk_email, chk_phone, chk_address, chk_social_links,
    chk_author, chk_date_published, chk_date_modified, chk_external_links,
    chk_toc, chk_references, chk_images, chk_comments, chk_tags,
    chk_reading_progress, chk_cookie, chk_reviews, chk_trustpilot,
    chk_licenses, chk_editor, chk_article_views, chk_article_rating_widget,
    chk_page_exists, chk_categories, chk_disclaimer, chk_callback,
    chk_online_chat, chk_search, chk_newsletter,
    chk_author_education, chk_author_experience,
    chk_org_mission, chk_org_age, chk_reg_docs, chk_team_photos, chk_form,
    get_evidence,
)
from recommendations import get_recommendation
from crawler import analyze_all_pages, PAGE_TYPE_LABELS, PAGE_CHECKS

SITE_TYPES = ["E-commerce", "Сайти послуг", "Блоги / сайти новин", "Аптеки / медицина"]
IMPORTANCE_ORDER = {"Висока": 0, "Середня": 1, "Низька": 2}
COLUMNS = ["Розділ", "Фактор", "Важливість", "Статус", "Рекомендація", "Приклад"]


# ─── ПОБУДОВА ЧЕКЛІСТУ ────────────────────────────────────────────────────────

def R(results, section, factor, importance, status, soup=None, schemas=None, base_url=""):
    """Додає рядок результату з автоматичною рекомендацією та прикладом."""
    rec = get_recommendation(factor) if status != OK else "—"
    example = get_evidence(factor, soup, schemas or [], base_url) if (status == OK and soup is not None) else ""
    results.append((section, factor, importance, status, rec, example))


def build_checklist(site_type, main_url, main_soup, about_soup, contact_soup, article_soup, author_soup):
    all_links    = get_internal_links(main_soup, main_url)
    main_schemas = get_schemas(main_soup)

    results  = []
    about_s  = about_soup   or main_soup
    contact_s = contact_soup or main_soup
    art      = article_soup or main_soup
    art_schemas = get_schemas(art)

    contact_schemas = get_schemas(contact_s)
    about_schemas   = get_schemas(about_s)
    auth_schemas    = []  # заповниться нижче

    # ── ЗАГАЛЬНЕ ────────────────────────────────────────────────────────────
    R(results, "Загальне", "HTTPS",              "Висока", chk_https(main_url),   main_soup, main_schemas, main_url)
    R(results, "Загальне", "Пошук по сайту",     "Низька",  chk_search(main_soup),    main_soup, main_schemas, main_url)
    R(results, "Загальне", "Плашка про куки",    "Низька",  chk_cookie(main_soup),    main_soup, main_schemas, main_url)
    R(results, "Загальне", "Онлайн-консультант", "Низька",  chk_online_chat(main_soup), main_soup, main_schemas, main_url)

    # ── СТОРІНКА «ПРО КОМПАНІЮ» ─────────────────────────────────────────────
    about_exists = OK if about_soup else chk_page_exists(
        all_links, ["про нас", "about", "про компанію", "o-nas", "/about"]
    )
    R(results, "Про компанію", "Сторінка існує", "Висока", about_exists, about_s, about_schemas, main_url)

    if site_type in ("Блоги / сайти новин", "Аптеки / медицина", "Сайти послуг"):
        R(results, "Про компанію", "Сфера діяльності компанії", "Висока",
          OK if about_s and len(about_s.get_text()) > 200 else WARN, about_s, about_schemas, main_url)
        R(results, "Про компанію", "Місія / цінності компанії", "Середня",   chk_org_mission(about_s),  about_s, about_schemas, main_url)
        R(results, "Про компанію", "Вік / дата заснування",     "Середня",   chk_org_age(about_s),      about_s, about_schemas, main_url)
        R(results, "Про компанію", "Фотографії команди",         "Низька",    chk_team_photos(about_s),  about_s, about_schemas, main_url)

    if site_type in ("Аптеки / медицина", "Сайти послуг", "Блоги / сайти новин"):
        imp = "Висока" if site_type == "Аптеки / медицина" else "Середня"
        R(results, "Про компанію", "Свідоцтво / реквізити юридичної особи", imp, chk_reg_docs(about_s), about_s, about_schemas, main_url)
        R(results, "Про компанію", "Нагороди та досягнення", "Середня",
          OK if about_s and any(w in about_s.get_text().lower() for w in ["нагород", "award", "перемог", "визнан"]) else FAIL,
          about_s, about_schemas, main_url)

    if site_type == "Блоги / сайти новин":
        R(results, "Про компанію", "Редакційна політика", "Висока",
          chk_page_exists(all_links, ["редакційна", "editorial policy", "редакційн"]), main_soup, main_schemas, main_url)
        R(results, "Про компанію", "Сторінка «Редакція» / «Команда»", "Висока",
          chk_page_exists(all_links, ["редакція", "команда", "/team", "/editorial"]), main_soup, main_schemas, main_url)

    if site_type == "Сайти послуг":
        R(results, "Про компанію", "ЗМІ про нас", "Середня",
          OK if about_s and any(w in about_s.get_text().lower() for w in ["зМІ", "преса", "media", "press", "згадки"]) else FAIL,
          about_s, about_schemas, main_url)

    # ── КОНТАКТНА СТОРІНКА ──────────────────────────────────────────────────
    contact_exists = OK if contact_soup else chk_page_exists(
        all_links, ["контакт", "contact", "зв'яжіться", "/contacts"]
    )
    R(results, "Контактна сторінка", "Сторінка існує",           "Висока", contact_exists,                    contact_s, contact_schemas, main_url)
    R(results, "Контактна сторінка", "Електронна адреса",         "Висока", chk_email(contact_s),              contact_s, contact_schemas, main_url)
    R(results, "Контактна сторінка", "Номер телефону",            "Висока", chk_phone(contact_s),              contact_s, contact_schemas, main_url)
    R(results, "Контактна сторінка", "Фізична адреса",            "Середня", chk_address(contact_s, contact_schemas), contact_s, contact_schemas, main_url)
    R(results, "Контактна сторінка", "Посилання на соцмережі",   "Середня", chk_social_links(main_soup),       main_soup, main_schemas, main_url)
    R(results, "Контактна сторінка", "Форма зворотного зв'язку", "Середня", chk_form(contact_s),               contact_s, contact_schemas, main_url)
    R(results, "Контактна сторінка", "Контакти керівника",        "Низька",
      OK if contact_s and any(w in contact_s.get_text().lower() for w in ["директор", "ceo", "керівник", "founder"]) else FAIL,
      contact_s, contact_schemas, main_url)

    if site_type == "Сайти послуг":
        R(results, "Контактна сторінка", "Замовлення зворотного дзвінка", "Низька", chk_callback(main_soup), main_soup, main_schemas, main_url)
        R(results, "Контактна сторінка", "Схема проїзду / карта", "Низька",
          OK if contact_s and any(w in str(contact_s).lower() for w in ["google.com/maps", "maps.app", "openstreetmap", "iframe"]) else FAIL,
          contact_s, contact_schemas, main_url)

    # ── ВАЖЛИВІ СТОРІНКИ ────────────────────────────────────────────────────
    R(results, "Важливі сторінки", "Політика конфіденційності", "Висока",
      chk_page_exists(all_links, ["конфіденційність", "privacy", "policy"]), main_soup, main_schemas, main_url)

    if site_type != "Блоги / сайти новин":
        R(results, "Важливі сторінки", "Користувацька угода / Публічна оферта", "Висока",
          chk_page_exists(all_links, ["угода", "оферта", "terms", "пропозиція"]), main_soup, main_schemas, main_url)
        R(results, "Важливі сторінки", "Гарантії / умови повернення", "Висока",
          chk_page_exists(all_links, ["гарантія", "повернення", "guarantee", "return"]), main_soup, main_schemas, main_url)

    if site_type in ("Сайти послуг", "E-commerce", "Аптеки / медицина"):
        R(results, "Важливі сторінки", "Доставка", "Висока",
          chk_page_exists(all_links, ["доставка", "delivery", "shipping"]), main_soup, main_schemas, main_url)
        R(results, "Важливі сторінки", "Способи оплати", "Висока",
          chk_page_exists(all_links, ["оплата", "payment", "способи оплати"]), main_soup, main_schemas, main_url)

    if site_type in ("Сайти послуг", "Аптеки / медицина"):
        R(results, "Важливі сторінки", "Ліцензії та сертифікати",             "Висока", chk_licenses(about_s),            about_s, about_schemas, main_url)
        R(results, "Важливі сторінки", "Відгуки на зовнішніх платформах",     "Висока", chk_trustpilot(main_soup),         main_soup, main_schemas, main_url)
        R(results, "Важливі сторінки", "Відгуки на сайті",                    "Висока", chk_reviews(main_soup, main_schemas), main_soup, main_schemas, main_url)
        R(results, "Важливі сторінки", "Редакційна / маркетингова політика",  "Середня",
          chk_page_exists(all_links, ["редакційна", "маркетингова", "editorial"]), main_soup, main_schemas, main_url)
        R(results, "Важливі сторінки", "Профілі у відгуковиках (TrustPilot і т.д.)", "Висока", chk_trustpilot(main_soup), main_soup, main_schemas, main_url)

    R(results, "Важливі сторінки", "FAQ сторінка", "Низька",
      chk_page_exists(all_links, ["faq", "питання", "запитання", "відповіді"]), main_soup, main_schemas, main_url)

    if site_type == "Блоги / сайти новин":
        R(results, "Важливі сторінки", "Підписка на розсилку", "Низька", chk_newsletter(main_soup), main_soup, main_schemas, main_url)

    # ── МІКРОРОЗМІТКА ───────────────────────────────────────────────────────
    R(results, "Мікророзмітка", "Organization", "Висока",
      OK if has_schema_type(main_schemas, "Organization") else FAIL, main_soup, main_schemas, main_url)
    breadcrumb_imp = "Висока" if site_type == "E-commerce" else "Середня"
    R(results, "Мікророзмітка", "BreadcrumbList", breadcrumb_imp,
      OK if has_schema_type(main_schemas, "BreadcrumbList") else FAIL, main_soup, main_schemas, main_url)
    R(results, "Мікророзмітка", "Article / BlogPosting", "Висока",
      OK if has_schema_type(main_schemas, "Article", "BlogPosting", "NewsArticle") else FAIL, main_soup, main_schemas, main_url)
    R(results, "Мікророзмітка", "FAQPage", "Середня",
      OK if has_schema_type(main_schemas, "FAQPage") else FAIL, main_soup, main_schemas, main_url)

    if site_type in ("E-commerce", "Аптеки / медицина"):
        R(results, "Мікророзмітка", "Product", "Висока",
          OK if has_schema_type(main_schemas, "Product") else FAIL, main_soup, main_schemas, main_url)
    if site_type in ("Блоги / сайти новин", "Аптеки / медицина"):
        R(results, "Мікророзмітка", "Person", "Висока",
          OK if has_schema_type(main_schemas, "Person") else FAIL, main_soup, main_schemas, main_url)
    if site_type == "Сайти послуг":
        R(results, "Мікророзмітка", "LocalBusiness", "Середня",
          OK if has_schema_type(main_schemas, "LocalBusiness") else FAIL, main_soup, main_schemas, main_url)
    if site_type == "Аптеки / медицина":
        R(results, "Мікророзмітка", "MedicalWebPage",   "Висока",
          OK if has_schema_type(main_schemas, "MedicalWebPage") else FAIL,   main_soup, main_schemas, main_url)
        R(results, "Мікророзмітка", "MedicalClinic",    "Низька",
          OK if has_schema_type(main_schemas, "MedicalClinic") else FAIL,    main_soup, main_schemas, main_url)
        R(results, "Мікророзмітка", "MedicalCondition", "Низька",
          OK if has_schema_type(main_schemas, "MedicalCondition") else FAIL, main_soup, main_schemas, main_url)

    # ── СТАТТІ ──────────────────────────────────────────────────────────────
    R(results, "Статті", "Зазначено автора",                 "Висока",  chk_author(art, art_schemas),          art, art_schemas, main_url)
    R(results, "Статті", "Дата публікації",                  "Висока",  chk_date_published(art, art_schemas),  art, art_schemas, main_url)
    R(results, "Статті", "Дата оновлення контенту",          "Висока",  chk_date_modified(art, art_schemas),   art, art_schemas, main_url)
    R(results, "Статті", "Посилання на авторитетні ресурси", "Висока",  chk_external_links(art, main_url),     art, art_schemas, main_url)
    R(results, "Статті", "Окремі сторінки для авторів",      "Висока",
      chk_page_exists(all_links, ["автор", "author", "/authors", "/team"]), main_soup, main_schemas, main_url)

    editor_imp = "Висока" if site_type == "Аптеки / медицина" else "Середня"
    R(results, "Статті", "Редактор / рецензент матеріалу",   editor_imp, chk_editor(art, art_schemas), art, art_schemas, main_url)

    toc_imp = "Висока" if site_type == "Аптеки / медицина" else "Середня"
    R(results, "Статті", "Зміст статті (Table of Contents)", toc_imp,   chk_toc(art), art, art_schemas, main_url)

    ref_imp = "Висока" if site_type == "Аптеки / медицина" else "Середня"
    R(results, "Статті", "Список літератури / джерела",      ref_imp,   chk_references(art), art, art_schemas, main_url)

    R(results, "Статті", "Розподіл статей на категорії",     "Середня", chk_categories(main_soup, all_links),    main_soup, main_schemas, main_url)
    R(results, "Статті", "Кількість переглядів",             "Середня", chk_article_views(art),                  art, art_schemas, main_url)
    R(results, "Статті", "Можливість оцінки статті",         "Середня", chk_article_rating_widget(art),          art, art_schemas, main_url)
    R(results, "Статті", "Наявність зображень",              "Середня", chk_images(art),                         art, art_schemas, main_url)
    R(results, "Статті", "Можливість залишити коментар",     "Середня", chk_comments(art),                       art, art_schemas, main_url)
    R(results, "Статті", "Теги та хмара тегів",              "Низька",  chk_tags(art),                           art, art_schemas, main_url)
    R(results, "Статті", "Повзунок прогресу читання",        "Низька",  chk_reading_progress(art),               art, art_schemas, main_url)

    if site_type == "Аптеки / медицина":
        R(results, "Статті", "Відмова від відповідальності", "Висока",  chk_disclaimer(art), art, art_schemas, main_url)

    # ── СТОРІНКИ АВТОРІВ ────────────────────────────────────────────────────
    auth = author_soup or main_soup
    auth_schemas = get_schemas(auth)

    R(results, "Сторінки авторів", "Окрема сторінка для автора",       "Висока",
      OK if author_soup else chk_page_exists(all_links, ["автор", "author", "/team", "/authors"]),
      auth, auth_schemas, main_url)
    R(results, "Сторінки авторів", "Зазначено освіту",                  "Висока", chk_author_education(auth),  auth, auth_schemas, main_url)
    R(results, "Сторінки авторів", "Зазначено досвід роботи",           "Висока", chk_author_experience(auth), auth, auth_schemas, main_url)
    R(results, "Сторінки авторів", "Посилання на соцмережі автора",    "Середня", chk_social_links(auth),      auth, auth_schemas, main_url)
    R(results, "Сторінки авторів", "Згадки автора на інших платформах","Середня", WARN)

    if site_type == "Аптеки / медицина":
        R(results, "Сторінки лікарів", "Окремі сторінки для лікарів", "Висока",
          chk_page_exists(all_links, ["лікар", "doctor", "physician", "спеціаліст"]), main_soup, main_schemas, main_url)
        R(results, "Сторінки лікарів", "Спеціальність лікаря", "Висока",
          OK if auth and any(w in auth.get_text().lower() for w in ["спеціальність", "specialty", "лікар"]) else FAIL,
          auth, auth_schemas, main_url)
        R(results, "Сторінки лікарів", "Медична освіта та кваліфікація", "Висока", chk_author_education(auth),  auth, auth_schemas, main_url)
        R(results, "Сторінки лікарів", "Стаж лікаря",                   "Висока", chk_author_experience(auth), auth, auth_schemas, main_url)
        R(results, "Сторінки лікарів", "Ліцензії та сертифікати лікаря","Висока", chk_licenses(auth),          auth, auth_schemas, main_url)
        R(results, "Сторінки лікарів", "Посилання на профілі в соцмережах", "Середня", chk_social_links(auth), auth, auth_schemas, main_url)

    return results


# ─── ЗВАЖЕНИЙ БАЛ ─────────────────────────────────────────────────────────────

def calc_score(df: pd.DataFrame) -> int:
    weights = {"Висока": 3, "Середня": 2, "Низька": 1}
    total_w = sum(weights.get(imp, 1) for imp in df["Важливість"])
    ok_w = sum(
        weights.get(row["Важливість"], 1)
        for _, row in df.iterrows()
        if OK in str(row["Статус"])
    )
    return int(ok_w / total_w * 100) if total_w else 0


# ─── EXCEL EXPORT ─────────────────────────────────────────────────────────────

def to_excel(df: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ── Аркуш 1: повні результати ──
        df.to_excel(writer, index=False, sheet_name="Повний аналіз")
        ws = writer.sheets["Повний аналіз"]

        col_widths = {"A": 22, "B": 46, "C": 12, "D": 28, "E": 70, "F": 50}
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w

        fill_ok   = PatternFill("solid", fgColor="C6EFCE")
        fill_fail = PatternFill("solid", fgColor="FFC7CE")
        fill_warn = PatternFill("solid", fgColor="FFEB9C")
        fill_head = PatternFill("solid", fgColor="4472C4")

        for cell in ws[1]:
            cell.fill = fill_head
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows(min_row=2):
            status = str(row[3].value or "")
            if   "✅" in status: fill = fill_ok
            elif "❌" in status: fill = fill_fail
            elif "⚠️" in status: fill = fill_warn
            else:                fill = None
            if fill:
                for cell in row:
                    cell.fill = fill
            # Wrap text for recommendation and example columns
            row[4].alignment = Alignment(wrap_text=True, vertical="top")
            row[5].alignment = Alignment(wrap_text=True, vertical="top")

        # ── Аркуш 2: лише провали + рекомендації ──
        fails_df = df[df["Статус"] != OK][["Розділ", "Фактор", "Важливість", "Рекомендація"]].copy()
        fails_df["_sort"] = fails_df["Важливість"].map({"Висока": 0, "Середня": 1, "Низька": 2})
        fails_df = fails_df.sort_values(["_sort", "Розділ"]).drop(columns="_sort")

        fails_df.to_excel(writer, index=False, sheet_name="Рекомендації")
        ws2 = writer.sheets["Рекомендації"]

        col_widths2 = {"A": 22, "B": 46, "C": 12, "D": 80}
        for col, w in col_widths2.items():
            ws2.column_dimensions[col].width = w

        for cell in ws2[1]:
            cell.fill = fill_head
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        fill_high = PatternFill("solid", fgColor="FFC7CE")
        fill_mid  = PatternFill("solid", fgColor="FFEB9C")
        fill_low  = PatternFill("solid", fgColor="DDEBF7")

        for row in ws2.iter_rows(min_row=2):
            imp = str(row[2].value or "")
            if   imp == "Висока":  fill = fill_high
            elif imp == "Середня": fill = fill_mid
            else:                  fill = fill_low
            for cell in row:
                cell.fill = fill
            row[3].alignment = Alignment(wrap_text=True, vertical="top")

    return output.getvalue()


# ─── АНАЛІЗ ───────────────────────────────────────────────────────────────────

def run_analysis(url: str, site_type: str, progress):
    if not url.startswith("http"):
        url = "https://" + url

    progress.progress(10, "Завантаження головної сторінки...")
    main_soup = fetch(url)
    if not main_soup:
        return None, url

    all_links = get_internal_links(main_soup, url)

    def try_fetch(keywords):
        link = find_page_url(all_links, keywords)
        if link and link != url:
            return fetch(link)
        return None

    progress.progress(25, "Пошук сторінки «Про нас»...")
    about_soup = try_fetch(["про нас", "about", "про компанію", "/about"])

    progress.progress(40, "Пошук контактної сторінки...")
    contact_soup = try_fetch(["контакт", "contact", "зв'яжіться", "/contacts"])

    progress.progress(55, "Пошук статті / блогу...")
    article_soup = try_fetch(["blog", "блог", "статт", "новин", "article", "/news"])

    progress.progress(70, "Пошук сторінки автора...")
    _raw_author = try_fetch(["автор", "author", "/authors", "/team"])
    # Якщо знайдена сторінка є каталогом авторів книг (e-commerce) — ігноруємо її
    from crawler import _is_book_or_product_author_page
    author_soup = None if _is_book_or_product_author_page(_raw_author) else _raw_author

    progress.progress(85, "Аналіз факторів E-E-A-T...")
    results = build_checklist(
        site_type, url,
        main_soup, about_soup, contact_soup, article_soup, author_soup,
    )

    progress.progress(100, "Готово!")
    return results, url


# ─── UI ───────────────────────────────────────────────────────────────────────

def show_metrics(df):
    total      = len(df)
    ok_count   = df["Статус"].str.contains("✅").sum()
    fail_count = df["Статус"].str.contains("❌").sum()
    warn_count = df["Статус"].str.contains("⚠️").sum()
    score      = calc_score(df)

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("🏆 Зважений бал",    f"{score}%")
    m2.metric("✅ Виконано",         ok_count)
    m3.metric("❌ Відсутнє",         fail_count)
    m4.metric("⚠️ Потребує уваги",  warn_count)
    m5.metric("📋 Всього факторів", total)


def tab_full_results(df):
    """Вкладка з повними результатами по розділах."""
    for section in df["Розділ"].unique():
        sec_df    = df[df["Розділ"] == section].copy()
        sec_ok    = sec_df["Статус"].str.contains("✅").sum()
        sec_total = len(sec_df)
        sec_score = int(sec_ok / sec_total * 100)

        emoji = "✅" if sec_score >= 75 else ("⚠️" if sec_score >= 40 else "❌")

        sec_df["_sort"] = sec_df["Важливість"].map(IMPORTANCE_ORDER)
        sec_df = sec_df.sort_values("_sort").drop(columns="_sort")

        with st.expander(
            f"{emoji} **{section}** — {sec_ok}/{sec_total} ({sec_score}%)",
            expanded=(sec_score < 75),
        ):
            st.dataframe(
                sec_df[["Фактор", "Важливість", "Статус", "Приклад"]],
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Фактор":     st.column_config.TextColumn("Фактор",                    width="large"),
                    "Важливість": st.column_config.TextColumn("Важливість",                width="small"),
                    "Статус":     st.column_config.TextColumn("Статус",                    width="medium"),
                    "Приклад":    st.column_config.TextColumn("Приклад (що знайдено)",     width="large"),
                },
            )


def tab_recommendations(df):
    """Вкладка: що потрібно виправити, згруповано по пріоритету."""
    fails = df[df["Статус"] != OK].copy()

    if fails.empty:
        st.success("🎉 Чудово! Всі фактори пройдені — рекомендацій немає.")
        return

    for imp_label, color, icon in [
        ("Висока",  "#ffc7ce", "🔴"),
        ("Середня", "#ffeb9c", "🟡"),
        ("Низька",  "#ddebf7", "🔵"),
    ]:
        group = fails[fails["Важливість"] == imp_label]
        if group.empty:
            continue

        st.markdown(f"### {icon} Важливість: {imp_label} ({len(group)} пунктів)")

        for _, row in group.iterrows():
            status_icon = "❌" if "❌" in str(row["Статус"]) else "⚠️"
            with st.expander(f"{status_icon} **{row['Розділ']}** → {row['Фактор']}"):
                st.markdown(f"**Що зробити:**")
                st.info(row["Рекомендація"])

        st.divider()


def tab_pages_analysis(crawl_data: dict):
    """Вкладка: детальний аналіз всіх знайдених сторінок."""
    if not crawl_data:
        st.info("Сторінки авторів/лікарів/статей не знайдено автоматично. "
                "Можливо, посилання на них схованні глибше в структурі сайту.")
        return

    for page_type, pages in crawl_data.items():
        label       = PAGE_TYPE_LABELS.get(page_type, page_type)
        check_names = [c[0] for c in PAGE_CHECKS.get(page_type, [])]
        total_pages = len(pages)
        ok_pages    = [p for p in pages if not p.get("error")]

        st.markdown(f"## {label}")
        st.caption(f"Знайдено сторінок: **{total_pages}**")

        if not ok_pages:
            st.warning("Не вдалося завантажити жодну сторінку цього типу.")
            continue

        # ── Зведена таблиця по факторах ──────────────────────────────────
        st.markdown("**Зведення по факторах:**")
        summary_rows = []
        for check_name in check_names:
            passed  = sum(1 for p in ok_pages if p["results"].get(check_name) == OK)
            total_c = len(ok_pages)
            pct     = int(passed / total_c * 100) if total_c else 0
            status  = "✅" if pct >= 80 else ("⚠️" if pct >= 40 else "❌")
            summary_rows.append({
                "Фактор":    check_name,
                "Виконано":  f"{passed} / {total_c}",
                "%":         f"{pct}%",
                "Статус":    status,
            })

        st.dataframe(
            pd.DataFrame(summary_rows),
            use_container_width=True,
            hide_index=True,
            column_config={
                "Фактор":   st.column_config.TextColumn(width="large"),
                "Виконано": st.column_config.TextColumn(width="small"),
                "%":        st.column_config.TextColumn(width="small"),
                "Статус":   st.column_config.TextColumn(width="small"),
            },
        )

        # ── Детальна таблиця по кожній сторінці ──────────────────────────
        with st.expander(f"📄 Детально по кожній сторінці ({total_pages})", expanded=False):
            rows = []
            for page in ok_pages:
                row = {"Сторінка": page["label"], "URL": page["url"]}
                row.update(page["results"])
                rows.append(row)

            if rows:
                detail_df = pd.DataFrame(rows)
                st.dataframe(
                    detail_df,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "URL": st.column_config.LinkColumn("URL", width="medium"),
                        "Сторінка": st.column_config.TextColumn(width="medium"),
                    },
                )

        # ── Сторінки з проблемами ─────────────────────────────────────────
        problem_pages = [
            p for p in ok_pages
            if any(v != OK for v in p["results"].values())
        ]
        if problem_pages:
            with st.expander(f"❌ Сторінки з відсутніми факторами ({len(problem_pages)})", expanded=True):
                for page in problem_pages:
                    missing = [k for k, v in page["results"].items() if v != OK]
                    st.markdown(
                        f"🔗 [{page['label']}]({page['url']})  \n"
                        f"Відсутнє: {', '.join(missing)}"
                    )

        st.divider()


def to_excel_with_crawl(df: pd.DataFrame, crawl_data: dict) -> bytes:
    """Excel з усіма аркушами: основний аналіз + окремий аркуш на кожен тип сторінок."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        fill_ok   = PatternFill("solid", fgColor="C6EFCE")
        fill_fail = PatternFill("solid", fgColor="FFC7CE")
        fill_warn = PatternFill("solid", fgColor="FFEB9C")
        fill_head = PatternFill("solid", fgColor="4472C4")

        def style_header(ws):
            for cell in ws[1]:
                cell.fill = fill_head
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")

        def status_fill(status_str):
            if "✅" in status_str:   return fill_ok
            if "❌" in status_str:   return fill_fail
            if "⚠️" in status_str:  return fill_warn
            return None

        # ── Аркуш 1: Повний аналіз ──────────────────────────────────────
        df.to_excel(writer, index=False, sheet_name="Повний аналіз")
        ws = writer.sheets["Повний аналіз"]
        for col, w in {"A": 22, "B": 46, "C": 12, "D": 28, "E": 70, "F": 50}.items():
            ws.column_dimensions[col].width = w
        style_header(ws)
        for row in ws.iter_rows(min_row=2):
            fill = status_fill(str(row[3].value or ""))
            if fill:
                for cell in row: cell.fill = fill
            row[4].alignment = Alignment(wrap_text=True, vertical="top")
            row[5].alignment = Alignment(wrap_text=True, vertical="top")

        # ── Аркуш 2: Рекомендації ────────────────────────────────────────
        fails_df = df[df["Статус"] != OK][["Розділ", "Фактор", "Важливість", "Рекомендація"]].copy()
        fails_df["_s"] = fails_df["Важливість"].map({"Висока": 0, "Середня": 1, "Низька": 2})
        fails_df = fails_df.sort_values(["_s", "Розділ"]).drop(columns="_s")
        fails_df.to_excel(writer, index=False, sheet_name="Рекомендації")
        ws2 = writer.sheets["Рекомендації"]
        for col, w in {"A": 22, "B": 46, "C": 12, "D": 80}.items():
            ws2.column_dimensions[col].width = w
        style_header(ws2)
        fill_hi  = PatternFill("solid", fgColor="FFC7CE")
        fill_mid = PatternFill("solid", fgColor="FFEB9C")
        fill_lo  = PatternFill("solid", fgColor="DDEBF7")
        for row in ws2.iter_rows(min_row=2):
            imp  = str(row[2].value or "")
            fill = fill_hi if imp == "Висока" else (fill_mid if imp == "Середня" else fill_lo)
            for cell in row: cell.fill = fill
            row[3].alignment = Alignment(wrap_text=True, vertical="top")

        # ── Аркуші по типах сторінок ─────────────────────────────────────
        for page_type, pages in crawl_data.items():
            ok_pages    = [p for p in pages if not p.get("error")]
            check_names = [c[0] for c in __import__("crawler").PAGE_CHECKS.get(page_type, [])]
            sheet_name  = PAGE_TYPE_LABELS.get(page_type, page_type)[:31]  # Excel limit

            rows = []
            for page in ok_pages:
                row = {"Назва": page["label"], "URL": page["url"]}
                row.update(page["results"])
                rows.append(row)

            if not rows:
                continue

            pd.DataFrame(rows).to_excel(writer, index=False, sheet_name=sheet_name)
            ws_p = writer.sheets[sheet_name]
            ws_p.column_dimensions["A"].width = 30
            ws_p.column_dimensions["B"].width = 50
            for i in range(len(check_names)):
                col_letter = chr(ord("C") + i)
                ws_p.column_dimensions[col_letter].width = 22
            style_header(ws_p)
            for row in ws_p.iter_rows(min_row=2):
                for cell in row:
                    fill = status_fill(str(cell.value or ""))
                    if fill:
                        cell.fill = fill

    return output.getvalue()


# ─── ЗАГЛУШКА ПРОГРЕСУ ────────────────────────────────────────────────────────

class _SilentProgress:
    """No-op progress for background competitor analysis."""
    def progress(self, pct, text=""): pass
    def empty(self): pass


# ─── АНАЛІЗ КОНКУРЕНТА ────────────────────────────────────────────────────────

def analyze_competitor(url: str, site_type: str):
    """Runs basic checklist for a competitor URL. Returns DataFrame or None."""
    results, _ = run_analysis(url.strip(), site_type, _SilentProgress())
    return pd.DataFrame(results, columns=COLUMNS) if results else None


# ─── ПОРІВНЯЛЬНА ТАБЛИЦЯ (ДАНІ) ───────────────────────────────────────────────

def build_comparison_df(main_df: pd.DataFrame, comp_dfs: list) -> pd.DataFrame:
    """
    comp_dfs: [(domain: str, df: pd.DataFrame | None), ...] — up to 4 pairs.
    Returns DataFrame with columns: Розділ, Фактор, Важливість, К1, К2, К3, К4, Сума, Наш сайт
    """
    rows = []
    for _, row in main_df.iterrows():
        factor = row["Фактор"]
        comp_vals = []
        for _, cdf in comp_dfs:
            if cdf is None:
                comp_vals.append("—")
            else:
                m = cdf[cdf["Фактор"] == factor]
                comp_vals.append(m.iloc[0]["Статус"] if len(m) else FAIL)
        while len(comp_vals) < 4:
            comp_vals.append("—")
        suma = sum(1 for v in comp_vals if OK in str(v))
        rows.append({
            "Розділ":     row["Розділ"],
            "Фактор":     factor,
            "Важливість": row["Важливість"],
            "К1":         comp_vals[0],
            "К2":         comp_vals[1],
            "К3":         comp_vals[2],
            "К4":         comp_vals[3],
            "Сума":       suma,
            "Наш сайт":  row["Статус"],
        })
    return pd.DataFrame(rows)


# ─── ВКЛАДКА «ПОРІВНЯННЯ» ─────────────────────────────────────────────────────

def show_comparison(comp_df: pd.DataFrame, comp_dfs: list, main_url: str, main_df: pd.DataFrame):
    """Renders the competitor comparison tab."""
    main_domain = urlparse(main_url).netloc.replace("www.", "")
    active  = [(d, cdf) for d, cdf in comp_dfs if cdf is not None]
    failed  = [d for d, cdf in comp_dfs if cdf is None and d]

    if failed:
        for d in failed:
            st.warning(f"Не вдалося проаналізувати: **{d}**")

    # Score comparison row
    score_cols = st.columns(1 + len(active))
    main_score = calc_score(main_df)
    score_cols[0].metric(f"🏆 {main_domain} (наш)", f"{main_score}%")
    for i, (domain, cdf) in enumerate(active):
        s = calc_score(cdf)
        delta = main_score - s
        score_cols[i + 1].metric(
            f"📊 {domain}",
            f"{s}%",
            delta=f"{delta:+}%",
            delta_color="inverse",
        )

    st.divider()

    # Build column rename map: К1→domain, etc.
    col_rename = {}
    for i, (domain, _) in enumerate(comp_dfs):
        col_rename[f"К{i+1}"] = domain if domain else f"Конкурент {i+1}"

    display_df = comp_df.rename(columns=col_rename)
    active_comp_cols = [col_rename[f"К{i+1}"] for i, (_, cdf) in enumerate(comp_dfs) if cdf is not None]
    show_cols = ["Фактор", "Важливість"] + active_comp_cols + ["Сума", "Наш сайт"]

    col_cfg = {
        "Фактор":     st.column_config.TextColumn("Фактор",     width="large"),
        "Важливість": st.column_config.TextColumn("Важливість", width="small"),
        "Сума":       st.column_config.NumberColumn("Сума",     width="small"),
        "Наш сайт":  st.column_config.TextColumn(main_domain,  width="medium"),
    }
    for c in active_comp_cols:
        col_cfg[c] = st.column_config.TextColumn(c, width="medium")

    for section in display_df["Розділ"].unique():
        sec   = display_df[display_df["Розділ"] == section]
        ok_n  = sec["Наш сайт"].str.contains("✅", na=False).sum()
        total = len(sec)
        pct   = int(ok_n / total * 100) if total else 0
        icon  = "✅" if pct >= 75 else ("⚠️" if pct >= 40 else "❌")

        with st.expander(
            f"{icon} **{section}** — {main_domain}: {ok_n}/{total} ({pct}%)",
            expanded=(pct < 75),
        ):
            st.dataframe(
                sec[show_cols].reset_index(drop=True),
                use_container_width=True,
                hide_index=True,
                column_config=col_cfg,
            )


# ─── EXCEL З ПОРІВНЯННЯМ ──────────────────────────────────────────────────────

def to_excel_with_comparison(
    main_df: pd.DataFrame,
    comp_dfs: list,
    crawl_data: dict,
    main_url: str,
) -> bytes:
    """
    Excel with comparison sheet (PDF-format) + existing full analysis sheets.
    comp_dfs: [(domain, df_or_None), ...]
    """
    output = BytesIO()
    comp_df = build_comparison_df(main_df, comp_dfs)
    main_domain = urlparse(main_url).netloc.replace("www.", "")

    # Build column names matching PDF template
    comp_col_map = {}  # К1→domain
    for i, (domain, _) in enumerate(comp_dfs):
        comp_col_map[f"К{i+1}"] = domain or f"Конкурент {i+1}"

    # Build Excel dataframe for comparison sheet
    xls_rows = []
    for _, row in comp_df.iterrows():
        xrow = {
            "Тип сайту / Розділ": row["Розділ"],
            "Фактор":              row["Фактор"],
            "Важливість":          row["Важливість"],
        }
        comp_sum = 0
        for ki in range(1, 5):
            col_name = comp_col_map.get(f"К{ki}", f"Конкурент {ki}")
            raw = row.get(f"К{ki}", "—")
            if OK in str(raw):
                xrow[col_name] = 1
                comp_sum += 1
            elif FAIL in str(raw) or WARN in str(raw):
                xrow[col_name] = 0
            else:
                xrow[col_name] = "—"
        xrow["Сума"] = comp_sum

        our_raw = row["Наш сайт"]
        xrow[main_domain] = 1 if OK in str(our_raw) else (0 if (FAIL in str(our_raw) or WARN in str(our_raw)) else "—")
        xrow["Коментарі"] = ""
        xls_rows.append(xrow)

    xls_df = pd.DataFrame(xls_rows)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        fill_ok   = PatternFill("solid", fgColor="C6EFCE")
        fill_fail = PatternFill("solid", fgColor="FFC7CE")
        fill_warn = PatternFill("solid", fgColor="FFEB9C")
        fill_head = PatternFill("solid", fgColor="4472C4")
        fill_sum  = PatternFill("solid", fgColor="D9E1F2")
        fill_hi   = PatternFill("solid", fgColor="FFC7CE")
        fill_mid  = PatternFill("solid", fgColor="FFEB9C")
        fill_lo   = PatternFill("solid", fgColor="DDEBF7")

        def style_header(ws):
            for cell in ws[1]:
                cell.fill = fill_head
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

        def status_fill(s):
            s = str(s or "")
            if "✅" in s: return fill_ok
            if "❌" in s: return fill_fail
            if "⚠️" in s: return fill_warn
            return None

        # ── Sheet 1: Порівняння ──────────────────────────────────────────
        xls_df.to_excel(writer, index=False, sheet_name="Порівняння")
        ws = writer.sheets["Порівняння"]
        style_header(ws)

        # Set column widths
        ws.column_dimensions["A"].width = 22  # Тип сайту
        ws.column_dimensions["B"].width = 46  # Фактор
        ws.column_dimensions["C"].width = 12  # Важливість
        for col_letter, w in zip("DEFG", [20, 20, 20, 20]):
            ws.column_dimensions[col_letter].width = w
        ws.column_dimensions["H"].width = 8   # Сума
        ws.column_dimensions["I"].width = 20  # Наш сайт
        ws.column_dimensions["J"].width = 40  # Коментарі

        # Identify which columns are competitor/our site columns by index
        header_vals = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        our_and_comp_domains = [comp_col_map.get(f"К{i+1}", f"Конкурент {i+1}") for i in range(4)] + [main_domain]

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = header_vals[cell.column - 1] if cell.column <= len(header_vals) else ""
                if header in our_and_comp_domains:
                    if cell.value == 1:
                        cell.fill = fill_ok
                        cell.value = "✅"
                        cell.alignment = Alignment(horizontal="center")
                    elif cell.value == 0:
                        cell.fill = fill_fail
                        cell.value = "❌"
                        cell.alignment = Alignment(horizontal="center")
                    else:
                        cell.alignment = Alignment(horizontal="center")
                elif header == "Сума":
                    cell.fill = fill_sum
                    cell.alignment = Alignment(horizontal="center")
                elif header == "Важливість":
                    imp = str(cell.value or "")
                    if   imp == "Висока":  cell.fill = fill_hi
                    elif imp == "Середня": cell.fill = fill_mid
                    elif imp == "Низька":  cell.fill = fill_lo

        # ── Sheet 2: Повний аналіз ───────────────────────────────────────
        main_df.to_excel(writer, index=False, sheet_name="Повний аналіз")
        ws2 = writer.sheets["Повний аналіз"]
        for col, w in {"A": 22, "B": 46, "C": 12, "D": 28, "E": 70, "F": 50}.items():
            ws2.column_dimensions[col].width = w
        style_header(ws2)
        for row in ws2.iter_rows(min_row=2):
            fill = status_fill(str(row[3].value or ""))
            if fill:
                for cell in row: cell.fill = fill
            row[4].alignment = Alignment(wrap_text=True, vertical="top")
            row[5].alignment = Alignment(wrap_text=True, vertical="top")

        # ── Sheet 3: Рекомендації ────────────────────────────────────────
        fails_df = main_df[main_df["Статус"] != OK][["Розділ", "Фактор", "Важливість", "Рекомендація"]].copy()
        fails_df["_s"] = fails_df["Важливість"].map({"Висока": 0, "Середня": 1, "Низька": 2})
        fails_df = fails_df.sort_values(["_s", "Розділ"]).drop(columns="_s")
        fails_df.to_excel(writer, index=False, sheet_name="Рекомендації")
        ws3 = writer.sheets["Рекомендації"]
        for col, w in {"A": 22, "B": 46, "C": 12, "D": 80}.items():
            ws3.column_dimensions[col].width = w
        style_header(ws3)
        for row in ws3.iter_rows(min_row=2):
            imp = str(row[2].value or "")
            fill = fill_hi if imp == "Висока" else (fill_mid if imp == "Середня" else fill_lo)
            for cell in row: cell.fill = fill
            row[3].alignment = Alignment(wrap_text=True, vertical="top")

        # ── Crawl sheets ─────────────────────────────────────────────────
        for page_type, pages in crawl_data.items():
            ok_pages    = [p for p in pages if not p.get("error")]
            check_names = [c[0] for c in __import__("crawler").PAGE_CHECKS.get(page_type, [])]
            sheet_name  = PAGE_TYPE_LABELS.get(page_type, page_type)[:31]
            rows_data = []
            for page in ok_pages:
                r = {"Назва": page["label"], "URL": page["url"]}
                r.update(page["results"])
                rows_data.append(r)
            if not rows_data:
                continue
            pd.DataFrame(rows_data).to_excel(writer, index=False, sheet_name=sheet_name)
            ws_p = writer.sheets[sheet_name]
            ws_p.column_dimensions["A"].width = 30
            ws_p.column_dimensions["B"].width = 50
            for idx in range(len(check_names)):
                col_letter = chr(ord("C") + idx)
                ws_p.column_dimensions[col_letter].width = 22
            style_header(ws_p)
            for row in ws_p.iter_rows(min_row=2):
                for cell in row:
                    f = status_fill(str(cell.value or ""))
                    if f: cell.fill = f

    return output.getvalue()


def main():
    st.set_page_config(
        page_title="E-E-A-T Checker",
        page_icon="🔍",
        layout="wide",
    )

    st.title("🔍 E-E-A-T Checker")
    st.caption(
        "Автоматична перевірка сайту за факторами "
        "**Experience · Expertise · Authoritativeness · Trustworthiness**"
    )

    # ── Поля конкурентів (поза формою — гарантовано рендеряться) ────────────────
    st.markdown("**Конкуренти** (необов'язково, від 1 до 4):")
    comp_cols_ui = st.columns(4)
    comp_inputs = []
    for i, col in enumerate(comp_cols_ui):
        with col:
            cu = st.text_input(
                f"Конкурент {i + 1}",
                placeholder=f"https://competitor{i + 1}.com",
                key=f"comp_{i}",
            )
            comp_inputs.append(cu.strip())

    # ── Основна форма ────────────────────────────────────────────────────────
    with st.form("check_form"):
        col1, col2, col3 = st.columns([4, 2, 1])
        with col1:
            url_input = st.text_input(
                "URL нашого сайту",
                placeholder="https://example.com",
                label_visibility="collapsed",
            )
        with col2:
            site_type = st.selectbox(
                "Тип сайту",
                SITE_TYPES,
                label_visibility="collapsed",
            )
        with col3:
            submitted = st.form_submit_button("▶ Перевірити", type="primary", use_container_width=True)

    if not submitted or not url_input:
        st.info("Введіть URL і натисніть «Перевірити»")
        return

    competitors = [u if u.startswith("http") else "https://" + u for u in comp_inputs if u]

    # ── Фаза 1: базовий аналіз головного сайту ──────────────────────────────
    progress = st.progress(0, "Починаємо аналіз...")
    results, final_url = run_analysis(url_input.strip(), site_type, progress)
    progress.empty()

    if not results:
        st.error(f"Не вдалося завантажити сайт: **{url_input}**\nПеревірте URL або спробуйте пізніше.")
        return

    df = pd.DataFrame(results, columns=COLUMNS)
    show_metrics(df)
    st.divider()

    # ── Фаза 2: аналіз конкурентів ──────────────────────────────────────────
    comp_dfs = []  # [(domain, df_or_None)]
    if competitors:
        comp_bar = st.progress(0, "Аналізуємо конкурентів...")
        for i, comp_url in enumerate(competitors):
            domain = urlparse(comp_url).netloc.replace("www.", "")
            pct = int(i / len(competitors) * 100)
            comp_bar.progress(pct, f"Аналізуємо {domain} ({i + 1}/{len(competitors)})...")
            cdf = analyze_competitor(comp_url, site_type)
            comp_dfs.append((domain, cdf))
        comp_bar.progress(100, "Готово!")
        comp_bar.empty()

    # Pad comp_dfs to always have 4 entries (with empty domain + None df)
    while len(comp_dfs) < 4:
        comp_dfs.append(("", None))

    # ── Фаза 3: краулінг головного сайту ────────────────────────────────────
    with st.spinner("🔍 Шукаємо та аналізуємо всі сторінки авторів / статей..."):
        crawl_progress = st.progress(0, "")
        main_soup = fetch(final_url)
        crawl_data = analyze_all_pages(
            base_url=final_url,
            main_soup=main_soup,
            site_type=site_type,
            max_per_type=25,
            progress_callback=lambda pct, msg: crawl_progress.progress(pct, msg),
        )
        crawl_progress.empty()

    # ── Вкладки ─────────────────────────────────────────────────────────────
    has_competitors = any(d for d, _ in comp_dfs)

    if has_competitors:
        tab1, tab2, tab3, tab4 = st.tabs([
            "🆚 Порівняння з конкурентами",
            "📋 Результати перевірки",
            "🔎 Аналіз всіх сторінок",
            "💡 Рекомендації",
        ])
        with tab1:
            show_comparison(
                build_comparison_df(df, comp_dfs),
                comp_dfs,
                final_url,
                df,
            )
        with tab2:
            tab_full_results(df)
        with tab3:
            tab_pages_analysis(crawl_data)
        with tab4:
            tab_recommendations(df)
    else:
        tab1, tab2, tab3 = st.tabs([
            "📋 Результати перевірки",
            "🔎 Аналіз всіх сторінок",
            "💡 Рекомендації",
        ])
        with tab1:
            tab_full_results(df)
        with tab2:
            tab_pages_analysis(crawl_data)
        with tab3:
            tab_recommendations(df)

    # ── Скачати Excel ────────────────────────────────────────────────────────
    st.divider()
    domain = urlparse(final_url).netloc.replace("www.", "")

    if has_competitors:
        excel_data = to_excel_with_comparison(df, comp_dfs, crawl_data, final_url)
    else:
        excel_data = to_excel_with_crawl(df, crawl_data)

    st.download_button(
        label="📥 Завантажити Excel",
        data=excel_data,
        file_name=f"eeat_{domain}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )


if __name__ == "__main__":
    main()
